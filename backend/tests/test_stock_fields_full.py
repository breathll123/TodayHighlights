"""遍历 A 股全量，请求每只股票的所有字段，输出到 Excel。"""
import time
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUTPUT = Path(__file__).resolve().parent / "stock_fields_full.xlsx"

# 所有可能相关的字段范围
FIELD_RANGES = [
    "f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13,f14,f15,f16,f17,f18,f19,f20",
    "f21,f22,f23,f24,f25,f26,f27,f28,f29,f30,f31,f32,f33,f34,f35,f36,f37,f38,f39,f40",
    "f41,f42,f43,f44,f45,f46,f47,f48,f49,f50,f51,f52,f53,f54,f55,f56,f57,f58,f59,f60",
    "f61,f62,f63,f64,f65,f66,f67,f68,f69,f70,f71,f72,f73,f74,f75,f76,f77,f78,f79,f80",
    "f81,f82,f83,f84,f85,f86,f87,f88,f89,f90,f91,f92,f93,f94,f95,f96,f97,f98,f99,f100",
    "f101,f102,f103,f104,f105,f106,f107,f108,f109,f110,f111,f112,f113,f114,f115,f116,f117,f118,f119,f120",
    "f121,f122,f123,f124,f125,f126,f127,f128,f129,f130,f131,f132,f133,f134,f135,f136,f137,f138,f139,f140",
    "f141,f142,f143,f144,f145,f146,f147,f148,f149,f150,f151,f152,f153,f154,f155,f156,f157,f158,f159,f160",
    "f161,f162,f163,f164,f165,f166,f167,f168,f169,f170,f171,f172,f173,f174,f175,f176,f177,f178,f179,f180",
    "f181,f182,f183,f184,f185,f186,f187,f188,f189,f190,f191,f192,f193,f194,f195,f196,f197,f198,f199,f200",
    "f201,f202,f203,f204,f205,f206,f207,f208,f209,f210,f211,f212,f213,f214,f215,f216,f217,f218,f219,f220",
    "f221,f222,f223,f224,f225,f226,f227,f228,f229,f230,f231,f232,f233,f234,f235,f236,f237,f238,f239,f240",
    "f241,f242,f243,f244,f245,f246,f247,f248,f249,f250,f251,f252,f253,f254,f255,f256,f257,f258,f259,f260",
]

# 已知字段含义（持续补充）
KNOWN_FIELDS = {
    "f2": "最新价", "f3": "涨跌幅%", "f4": "涨跌额", "f5": "成交量(手)", "f6": "成交额",
    "f7": "振幅%", "f8": "换手率%", "f9": "市盈率(动态)", "f10": "量比",
    "f11": "5分钟涨跌", "f12": "股票代码", "f13": "市场", "f14": "股票名称",
    "f15": "最高", "f16": "最低", "f17": "今开", "f18": "昨收",
    "f20": "总市值", "f21": "流通市值", "f22": "涨速", "f23": "市净率",
    "f24": "60日涨跌幅%", "f25": "5日涨跌幅%", "f26": "上市日期",
    "f37": "加权量(手)", "f38": "加权均价",
    "f43": "最新价(千分)", "f44": "最高(千分)", "f45": "最低(千分)", "f46": "今开(千分)",
    "f47": "成交量(手)", "f48": "成交额(元)", "f49": "外盘", "f50": "内盘",
    "f57": "股票代码", "f58": "股票名称", "f60": "昨收(千分)",
    "f62": "主力净流入", "f64": "超大单净流入", "f66": "大单净流入",
    "f100": "行业板块", "f102": "地域板块",
    "f115": "市盈率(静态)", "f116": "总股本", "f117": "流通股本",
    "f127": "概念板块1", "f128": "概念板块2",
    "f136": "市场类型", "f137": "市场代码",
    "f152": "龙虎榜标记", "f153": "龙虎榜次数",
    "f161": "机构评级", "f162": "评级机构数", "f163": "目标价",
    "f169": "涨跌额(千分)", "f170": "涨跌幅%(千分)",
    "f174": "买入金额", "f175": "买入占比", "f176": "卖出金额", "f177": "卖出占比",
    "f178": "龙虎榜JSON", "f179": "龙虎榜JSON2",
    "f184": "数值1", "f185": "数值2", "f186": "数值3", "f187": "数值4", "f188": "数值5",
    "f189": "上市日期(数字)", "f190": "数值6", "f191": "数值7", "f192": "数值8", "f193": "数值9",
    "f250": "数值A", "f251": "数值B", "f252": "数值C", "f253": "数值D", "f254": "数值E",
    "f255": "数值F", "f256": "数值G", "f257": "数值H", "f258": "数值I", "f259": "数值J",
}

HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": "https://data.eastmoney.com/"}


def fetch_all_stocks():
    """Get all A-share stock codes via clist."""
    codes = []
    for p in range(1, 60):
        resp = httpx.get(
            "https://push2delay.eastmoney.com/api/qt/clist/get",
            params={"pn": p, "pz": 100, "po": 1, "np": 1, "fltt": 2, "invt": 2, "fid": "f12",
                    "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
                    "fields": "f12,f14"},
            headers=HEADERS, timeout=15,
        )
        data = resp.json().get("data")
        if data is None:
            break
        items = data.get("diff") or []
        if not items:
            break
        codes.extend((i["f12"], i.get("f14", "?")) for i in items)
        print(f"  page {p}: {len(items)} stocks  total={len(codes)}")
        time.sleep(0.1)
    print(f"\n共 {len(codes)} 只股票")
    return codes


def main():
    print("Step 1: 拉取全部 A 股代码...")
    stocks = fetch_all_stocks()

    # Process ALL stocks
    sample = stocks
    print(f"Step 2: 请求全部 {len(sample)} 只股票的详细字段...")

    # Build full fields string
    all_field_keys = []
    for r in FIELD_RANGES:
        all_field_keys.extend(r.split(","))

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Fields"

    # Header row
    ws.cell(row=1, column=1, value="股票代码").font = Font(bold=True)
    ws.cell(row=1, column=2, value="股票名称").font = Font(bold=True)
    for ci, fk in enumerate(all_field_keys, start=3):
        label = KNOWN_FIELDS.get(fk, fk)
        cell = ws.cell(row=1, column=ci, value=f"{fk} ({label})")
        cell.font = Font(bold=True, size=8)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c in range(1, len(all_field_keys) + 3):
        ws.cell(row=1, column=c).fill = header_fill

    row = 2
    for code, name in sample:
        # Build secid
        if code.startswith("6"):
            secid = f"1.{code}"
        else:
            secid = f"0.{code}"

        try:
            resp = httpx.get(
                "https://push2delay.eastmoney.com/api/qt/stock/get",
                params={"secid": secid, "fields": ",".join(all_field_keys)},
                headers=HEADERS, timeout=15,
            )
            data = resp.json().get("data", {})

            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=name)

            for ci, fk in enumerate(all_field_keys, start=3):
                val = data.get(fk, "")
                if val is None:
                    val = ""
                ws.cell(row=row, column=ci, value=str(val))

            print(f"  [{row-1}/{len(sample)}] {name}({code}) — {len([v for v in data.values() if v and v!='-' and v!=''])} non-empty fields")
        except Exception as e:
            print(f"  [{row-1}/{len(sample)}] {name}({code}) — ERROR: {e}")
            ws.cell(row=row, column=2, value=f"{name} (ERROR: {e})")

        row += 1
        time.sleep(0.1)

    # Freeze header row
    ws.freeze_panes = "A2"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n已写入: {OUTPUT}")


if __name__ == "__main__":
    main()
